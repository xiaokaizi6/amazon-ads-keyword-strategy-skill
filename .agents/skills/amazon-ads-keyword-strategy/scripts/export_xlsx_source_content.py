#!/usr/bin/env python3
"""Export every non-empty XLSX cell as source-faithful searchable JSONL."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--source-id", required=True)
    parser.add_argument("--reviewed-at", default=date.today().isoformat())
    args = parser.parse_args()

    workbook = load_workbook(args.source, data_only=False, read_only=True)
    source_hash = sha256(args.source)
    count = 0
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8", newline="\n") as handle:
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    value = cell.value
                    record = {
                        "record_id": f"{args.source_id}-XLSX-{sheet.title}-{cell.coordinate}",
                        "source_id": args.source_id,
                        "source_location": f"XLSX worksheet {sheet.title!r} cell {cell.coordinate}",
                        "content_type": "formula" if cell.data_type == "f" else "cell",
                        "content": str(value),
                        "cell_value": str(value),
                        "cell_data_type": cell.data_type,
                        "formula": str(value) if cell.data_type == "f" else None,
                        "number_format": cell.number_format,
                        "source_sha256": source_hash,
                        "reviewed_at": args.reviewed_at,
                    }
                    handle.write(json.dumps(record, ensure_ascii=False, separators=(",", ":")) + "\n")
                    count += 1
    print(f"Exported {count} non-empty XLSX cells to {args.output}")


if __name__ == "__main__":
    main()
